
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, F, Q
from django.utils import timezone
from datetime import datetime, timedelta
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
import io
import csv
from sales.models import Sale, SaleItem
from products.models import Product
from customers.models import Customer
from finances.models import Transaction
from .models import Report
import json

# Importación condicional de openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False




@api_view(['GET'])
@permission_classes([AllowAny])
def get_report_types(request):
    """Obtener tipos de reportes disponibles"""
    report_types = [
        {
            'key': 'sales',
            'name': 'Reporte de Ventas',
            'description': 'Analiza el rendimiento de ventas por período',
            'icon': '📊'
        },
        {
            'key': 'inventory',
            'name': 'Reporte de Inventario',
            'description': 'Estado actual del inventario y movimientos',
            'icon': '📦'
        },
        {
            'key': 'financial',
            'name': 'Reporte Financiero',
            'description': 'Ingresos, gastos y rentabilidad del negocio',
            'icon': '💰'
        },
        {
            'key': 'customers',
            'name': 'Reporte de Clientes',
            'description': 'Análisis de comportamiento de clientes',
            'icon': '👥'
        }
    ]
    return Response(report_types)



@api_view(['GET'])
@permission_classes([AllowAny])  # Permitir acceso sin autenticación para pruebas
def generate_sales_report(request):
    """Generar reporte de ventas"""
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        format_type = request.GET.get('format', 'json')  # json, csv, excel
        
        # Filtros de fecha
        filter_kwargs = {}
        if start_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            filter_kwargs['created_at__gte'] = timezone.make_aware(start_dt, timezone.get_current_timezone())
        if end_date:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            filter_kwargs['created_at__lte'] = timezone.make_aware(end_dt, timezone.get_current_timezone())
        
        # Obtener datos de ventas
        sales = Sale.objects.filter(**filter_kwargs).select_related('customer').order_by('-created_at')
        
        # Estadísticas generales
        total_sales = sales.aggregate(
            total_amount=Sum('total_amount'),
            total_count=Count('id')
        )
        
        # Datos por producto
        products_data = SaleItem.objects.filter(
            sale__in=sales
        ).values(
            'product__name', 'product__sku'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('unit_price')),
            sales_count=Count('sale', distinct=True)
        ).order_by('-total_revenue')
        
        # Datos por cliente
        customers_data = sales.values(
            'customer__name', 'customer__email'
        ).annotate(
            total_spent=Sum('total_amount'),
            purchase_count=Count('id')
        ).order_by('-total_spent')
        
        # Datos por día
        daily_sales = sales.extra(
            select={'day': 'DATE(created_at)'}
        ).values('day').annotate(
            daily_total=Sum('total_amount'),
            daily_count=Count('id')
        ).order_by('day')
        
        # Serializar correctamente los valores numéricos
        def safe_float(val):
            try:
                return float(val)
            except Exception:
                return 0.0

        top_products = []
        for prod in products_data[:10]:
            prod = dict(prod)
            prod['total_quantity'] = safe_float(prod.get('total_quantity', 0))
            prod['total_revenue'] = safe_float(prod.get('total_revenue', 0))
            prod['sales_count'] = safe_float(prod.get('sales_count', 0))
            # Asegurar que no haya NaN
            for k in ['total_quantity', 'total_revenue', 'sales_count']:
                if prod[k] is None or prod[k] != prod[k]:
                    prod[k] = 0.0
            top_products.append(prod)

        top_customers = []
        for cust in customers_data[:10]:
            cust = dict(cust)
            cust['total_spent'] = safe_float(cust.get('total_spent', 0))
            cust['purchase_count'] = safe_float(cust.get('purchase_count', 0))
            for k in ['total_spent', 'purchase_count']:
                if cust[k] is None or cust[k] != cust[k]:
                    cust[k] = 0.0
            top_customers.append(cust)

        daily_sales_serialized = []
        for day in daily_sales:
            day = dict(day)
            day['daily_total'] = safe_float(day.get('daily_total', 0))
            day['daily_count'] = safe_float(day.get('daily_count', 0))
            for k in ['daily_total', 'daily_count']:
                if day[k] is None or day[k] != day[k]:
                    day[k] = 0.0
            daily_sales_serialized.append(day)

        report_data = {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_amount': float(total_sales['total_amount'] or 0),
                'total_sales': total_sales['total_count'],
                'average_sale': float((total_sales['total_amount'] or 0) / max(total_sales['total_count'], 1))
            },
            'top_products': top_products,
            'top_customers': top_customers,
            'daily_sales': daily_sales_serialized,
            'generated_at': timezone.now().isoformat()
        }
        
        if format_type == 'json':
            return Response(report_data)
        elif format_type == 'csv':
            return generate_csv_response(report_data, 'sales_report')
        elif format_type == 'excel':
            return generate_excel_response(report_data, 'sales_report')
            
    except Exception as e:
        return Response(
            {'error': f'Error generando reporte: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])  # Permitir acceso sin autenticación para pruebas
def generate_inventory_report(request):
    """Generar reporte de inventario"""
    try:
        format_type = request.GET.get('format', 'json')
        
        # Obtener productos con stock bajo
        low_stock_products = Product.objects.filter(
            stock_quantity__lte=F('min_stock_level')
        ).values(
            'name', 'sku', 'stock_quantity', 'min_stock_level', 'price', 'cost'
        )
        
        # Estadísticas de inventario
        inventory_stats = Product.objects.aggregate(
            total_products=Count('id'),
            active_products=Count('id', filter=Q(is_active=True)),
            total_stock_value=Sum(F('stock_quantity') * F('cost')),
            low_stock_count=Count('id', filter=Q(stock_quantity__lte=F('min_stock_level')))
        )
        
        # Productos por categoría
        category_stats = Product.objects.values(
            'category__name'
        ).annotate(
            product_count=Count('id'),
            total_stock=Sum('stock_quantity'),
            total_value=Sum(F('stock_quantity') * F('cost'))
        ).order_by('-product_count')
        
        # Serializar correctamente los valores numéricos en low_stock_products
        low_stock_serialized = []
        for prod in low_stock_products:
            prod = dict(prod)
            for k in ['stock_quantity', 'min_stock_level', 'price', 'cost']:
                prod[k] = float(prod.get(k, 0) or 0)
                if prod[k] is None or prod[k] != prod[k]:
                    prod[k] = 0.0
            low_stock_serialized.append(prod)

        # Serializar correctamente los valores numéricos en category_breakdown
        category_serialized = []
        for cat in category_stats:
            cat = dict(cat)
            for k in ['product_count', 'total_stock', 'total_value']:
                cat[k] = float(cat.get(k, 0) or 0)
                if cat[k] is None or cat[k] != cat[k]:
                    cat[k] = 0.0
            category_serialized.append(cat)

        report_data = {
            'summary': {
                'total_products': inventory_stats['total_products'],
                'active_products': inventory_stats['active_products'],
                'total_stock_value': float(inventory_stats['total_stock_value'] or 0),
                'low_stock_count': inventory_stats['low_stock_count']
            },
            'low_stock_products': low_stock_serialized,
            'category_breakdown': category_serialized,
            'generated_at': timezone.now().isoformat()
        }
        
        if format_type == 'json':
            return Response(report_data)
        elif format_type == 'csv':
            return generate_csv_response(report_data, 'inventory_report')
        elif format_type == 'excel':
            return generate_excel_response(report_data, 'inventory_report')
            
    except Exception as e:
        return Response(
            {'error': f'Error generando reporte de inventario: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])  # Permitir acceso sin autenticación para pruebas
def generate_financial_report(request):
    """Generar reporte financiero"""
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        format_type = request.GET.get('format', 'json')
        
        # Filtros de fecha
        filter_kwargs = {}
        if start_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            filter_kwargs['transaction_date__gte'] = start_dt.date()
        if end_date:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            filter_kwargs['transaction_date__lte'] = end_dt.date()
        
        # Obtener transacciones
        transactions = Transaction.objects.filter(**filter_kwargs)
        
        # Ingresos y gastos
        income = transactions.filter(transaction_type='income').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        expenses = transactions.filter(transaction_type='expense').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        # Gastos por categoría
        expense_categories = transactions.filter(
            transaction_type='expense'
        ).values(
            'category__name'
        ).annotate(
            total_amount=Sum('amount'),
            transaction_count=Count('id')
        ).order_by('-total_amount')
        
        # Transacciones por mes (compatible con SQLite)
        monthly_data = transactions.extra(
            select={'month': 'strftime("%Y-%m", transaction_date)'}
        ).values('month', 'transaction_type').annotate(
            total_amount=Sum('amount')
        ).order_by('month')
        
        # Serializar correctamente los valores numéricos en expense_categories
        expense_serialized = []
        for cat in expense_categories:
            cat = dict(cat)
            cat['total_amount'] = float(cat.get('total_amount', 0) or 0)
            cat['transaction_count'] = float(cat.get('transaction_count', 0) or 0)
            if cat['total_amount'] is None or cat['total_amount'] != cat['total_amount']:
                cat['total_amount'] = 0.0
            if cat['transaction_count'] is None or cat['transaction_count'] != cat['transaction_count']:
                cat['transaction_count'] = 0.0
            expense_serialized.append(cat)

        # Serializar correctamente los valores numéricos en monthly_data
        monthly_serialized = []
        for month in monthly_data:
            month = dict(month)
            month['total_amount'] = float(month.get('total_amount', 0) or 0)
            if month['total_amount'] is None or month['total_amount'] != month['total_amount']:
                month['total_amount'] = 0.0
            monthly_serialized.append(month)

        report_data = {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_income': float(income),
                'total_expenses': float(expenses),
                'net_profit': float(income - expenses),
                'profit_margin': float((income - expenses) / max(income, 1) * 100)
            },
            'expense_categories': expense_serialized,
            'monthly_data': monthly_serialized,
            'generated_at': timezone.now().isoformat()
        }
        
        if format_type == 'json':
            return Response(report_data)
        elif format_type == 'csv':
            return generate_csv_response(report_data, 'financial_report')
        elif format_type == 'excel':
            return generate_excel_response(report_data, 'financial_report')
            
    except Exception as e:
        return Response(
            {'error': f'Error generando reporte financiero: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def generate_csv_response(data, filename):
    """Generar respuesta CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)

    # Escribir encabezados y datos según el tipo de reporte
    if 'summary' in data:
        writer.writerow(['Resumen del Reporte'])
        for key, value in data['summary'].items():
            writer.writerow([key.replace('_', ' ').title(), value])
        writer.writerow([])

    # Exportar top_products si existe
    if 'top_products' in data and data['top_products']:
        writer.writerow(['Top Productos'])
        headers = list(data['top_products'][0].keys())
        writer.writerow(headers)
        for prod in data['top_products']:
            writer.writerow([prod.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar top_customers si existe
    if 'top_customers' in data and data['top_customers']:
        writer.writerow(['Top Clientes'])
        headers = list(data['top_customers'][0].keys())
        writer.writerow(headers)
        for cust in data['top_customers']:
            writer.writerow([cust.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar daily_sales si existe
    if 'daily_sales' in data and data['daily_sales']:
        writer.writerow(['Ventas Diarias'])
        headers = list(data['daily_sales'][0].keys())
        writer.writerow(headers)
        for day in data['daily_sales']:
            writer.writerow([day.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar low_stock_products si existe
    if 'low_stock_products' in data and data['low_stock_products']:
        writer.writerow(['Productos con Stock Bajo'])
        headers = list(data['low_stock_products'][0].keys())
        writer.writerow(headers)
        for prod in data['low_stock_products']:
            writer.writerow([prod.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar category_breakdown si existe
    if 'category_breakdown' in data and data['category_breakdown']:
        writer.writerow(['Resumen por Categoría'])
        headers = list(data['category_breakdown'][0].keys())
        writer.writerow(headers)
        for cat in data['category_breakdown']:
            writer.writerow([cat.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar expense_categories si existe
    if 'expense_categories' in data and data['expense_categories']:
        writer.writerow(['Gastos por Categoría'])
        headers = list(data['expense_categories'][0].keys())
        writer.writerow(headers)
        for cat in data['expense_categories']:
            writer.writerow([cat.get(h, '') for h in headers])
        writer.writerow([])

    # Exportar monthly_data si existe
    if 'monthly_data' in data and data['monthly_data']:
        writer.writerow(['Transacciones Mensuales'])
        headers = list(data['monthly_data'][0].keys())
        writer.writerow(headers)
        for month in data['monthly_data']:
            writer.writerow([month.get(h, '') for h in headers])
        writer.writerow([])

    return response


def generate_excel_response(data, filename):
    """Generar respuesta Excel"""
    if not OPENPYXL_AVAILABLE:
        # Si openpyxl no está disponible, generar CSV en su lugar
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        if 'summary' in data:
            writer.writerow(['Resumen del Reporte'])
            for key, value in data['summary'].items():
                writer.writerow([key.replace('_', ' ').title(), value])
        
        return response
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="Reporte")
    else:
        ws.title = "Reporte"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    row = 1
    # Resumen
    if 'summary' in data:
        cell = ws.cell(row=row, column=1, value="Resumen del Reporte")
        cell.font = header_font
        cell.fill = header_fill
        row += 1
        for key, value in data['summary'].items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

    # Exportar bloques de datos
    def write_block(title, items):
        nonlocal row
        if items:
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = header_font
            cell.fill = header_fill
            row += 1
            headers = list(items[0].keys())
            for col_num, h in enumerate(headers, 1):
                ws.cell(row=row, column=col_num, value=h)
            row += 1
            for item in items:
                for col_num, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col_num, value=item.get(h, ''))
                row += 1
            row += 1

    write_block("Top Productos", data.get('top_products', []))
    write_block("Top Clientes", data.get('top_customers', []))
    write_block("Ventas Diarias", data.get('daily_sales', []))
    write_block("Productos con Stock Bajo", data.get('low_stock_products', []))
    write_block("Resumen por Categoría", data.get('category_breakdown', []))
    write_block("Gastos por Categoría", data.get('expense_categories', []))
    write_block("Transacciones Mensuales", data.get('monthly_data', []))

    # Ajustar ancho de columnas
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response
